import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def create_workbook():
    base_dir = "/mnt/d/SIH/gate1_benchmark"
    workbook_path = os.path.join(base_dir, "gate3_reviewer_workbook.xlsx")
    
    with open(os.path.join(base_dir, "gate3_human_validation.json"), "r", encoding="utf-8") as f:
        gate3 = json.load(f)
        
    with open(os.path.join(base_dir, "gate2_dataset.json"), "r", encoding="utf-8") as f:
        dataset = {item["id"]: item for item in json.load(f)}

    wb = openpyxl.Workbook()
    
    # 1. Instructions
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    instructions = [
        ["GATE 3: EduSetu Human Review Guide"],
        [],
        ["PASS = correct"],
        ["FAIL = incorrect"],
        ["UNCERTAIN = requires discussion/reference"],
        [],
        ["CRITICAL RULES:"],
        ["A grammatically fluent translation with scientifically incorrect terminology is a FAIL."],
        [],
        ["Examples:"],
        ["quadratic equation → quadrilateral", "terminology/semantic FAIL"],
        ["E = mc² → E = mc2", "formula-formatting FAIL"],
        ["F = ma → unrelated text", "critical semantic FAIL"],
        ["Python → translated/transliterated incorrectly", "technical identifier FAIL"],
        ["Detached Kannada suffix", "morphology/grammar FAIL"],
        [],
        ["Do not judge a translation against the AI reference alone."],
        ["Reviewers must use their linguistic and educational knowledge."],
        ["Leave all judgment cells empty if unreviewed. Fill with PASS/FAIL/UNCERTAIN."]
    ]
    for row in instructions:
        ws_instr.append(row)
        
    # 2. Hindi Review
    ws_hi = wb.create_sheet("Hindi Review")
    hi_headers = ["ID", "Domain", "Source English", "Raw Hindi", "Protected Hindi", "AI Reference Hindi",
                  "Semantic Correct", "Terminology Correct", "Grammar Correct", "Natural Fluency",
                  "Formula Correct", "Technical Identifier Correct", "Hallucination", "Omission", "Addition",
                  "Overall Verdict", "Reviewer Notes"]
    ws_hi.append(hi_headers)
    
    for item in gate3:
        dom = dataset[item["id"]]["domain"]
        row = [
            item["id"], dom, item["source_en"], item["raw_hi"], item["protected_hi"], item["ai_reference_hi"],
            "", "", "", "", "", "", "", "", "", "", ""
        ]
        ws_hi.append(row)
        
    # 3. Kannada Review
    ws_kn = wb.create_sheet("Kannada Review")
    kn_headers = ["ID", "Domain", "Source English", "Raw Kannada", "Protected Kannada", "Protected+Morphology Kannada",
                  "AI Reference Kannada", "Semantic Correct", "Terminology Correct", "Grammar Correct", "Natural Fluency",
                  "Morphology Correct", "Formula Correct", "Technical Identifier Correct", "Hallucination", "Omission", 
                  "Addition", "Overall Verdict", "Reviewer Notes"]
    ws_kn.append(kn_headers)
    
    for item in gate3:
        dom = dataset[item["id"]]["domain"]
        row = [
            item["id"], dom, item["source_en"], item["raw_kn"], item["protected_kn"], item["protected_morph_kn"], 
            item["ai_reference_kn"],
            "", "", "", "", "", "", "", "", "", "", "", ""
        ]
        ws_kn.append(row)
        
    # 4. Critical Failures (Placeholder for reviewers to log top issues)
    ws_crit = wb.create_sheet("Critical Failures")
    ws_crit.append(["ID", "Language", "Failure Type", "Description"])
    
    # 5. Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Metric", "Hindi Score", "Kannada Score"])
    
    wb.save(workbook_path)
    print("Workbook created.")

if __name__ == "__main__":
    create_workbook()
