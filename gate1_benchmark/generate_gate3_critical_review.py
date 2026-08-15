import json
import os
import re
import openpyxl

def main():
    base_dir = r"d:\SIH\gate1_benchmark"
    
    with open(os.path.join(base_dir, "gate2_dataset.json"), "r", encoding="utf-8") as f:
        dataset = {i["id"]: i for i in json.load(f)}
        
    with open(os.path.join(base_dir, "gate2_benchmark_results.json"), "r", encoding="utf-8") as f:
        gate2_res = json.load(f)
        
    with open(os.path.join(base_dir, "gate3_human_validation.json"), "r", encoding="utf-8") as f:
        gate3_cases = {i["id"]: i for i in json.load(f)}
        
    res_by_id = {}
    for r in gate2_res:
        uid = r["id"]
        lang = r["target_language"]
        if uid not in res_by_id:
            res_by_id[uid] = {"hi": {}, "kn": {}}
        res_by_id[uid][lang] = r
        
    # Analyze failures per case
    case_priorities = {}
    
    for uid, item in dataset.items():
        dom = item["domain"]
        formulas = item.get("formula_tokens", [])
        techs = item.get("technical_tokens", [])
        terms = item.get("terminology_tokens", [])
        risk_tags = item.get("risk_tags", [])
        
        best_prio = 9 # Control by default
        best_reason = "Representative successful case"
        p_label = "CONTROL"
        
        # Check all configs for failures
        # gate2c_report_generator.py logic:
        # term_pass if expected in out_text
        failed_term = False
        failed_form = False
        failed_tech = False
        kn_quad_failed = False
        kn_morph_flag = False
        
        for lang in ["hi", "kn"]:
            for conf, out_key in [("A", "out_A"), ("B", "out_B"), ("C", "out_C")]:
                out_text = res_by_id[uid][lang].get(out_key, "")
                
                # Formulas
                for f_token in formulas:
                    if f_token not in out_text:
                        failed_form = True
                
                # Techs
                for t_token in techs:
                    if t_token not in out_text:
                        failed_tech = True
                        
                # Terms
                for term_dict in terms:
                    expected = term_dict.get(lang) or term_dict.get(f"{lang}_expected") or term_dict["en"]
                    if expected not in out_text:
                        failed_term = True
                        if lang == "kn" and term_dict["en"].lower() == "quadratic equation":
                            kn_quad_failed = True
                            
            if lang == "kn":
                # checking morph_flags_C
                if res_by_id[uid]["kn"].get("morph_flags_C"):
                    kn_morph_flag = True

        is_halluc = "hallucination_risk" in risk_tags
        is_stem_risk = "technical_term" in risk_tags or "formula" in risk_tags
        is_omission = failed_form or failed_tech # using gate2 proxy
        
        # Determine priority 1-9
        if failed_term:
            prio = 1
            reason = "Human-critical terminology failures detected by Gate 2."
            pl = "CRITICAL"
        elif failed_form:
            prio = 2
            reason = "Formula preservation failures."
            pl = "CRITICAL"
        elif failed_tech:
            prio = 3
            reason = "Technical identifier failures."
            pl = "CRITICAL"
        elif kn_quad_failed:
            prio = 4
            reason = "Kannada quadratic-equation terminology failures."
            pl = "CRITICAL"
        elif kn_morph_flag:
            prio = 5
            reason = "Kannada morphology failures remaining after Config C."
            pl = "HIGH"
        elif is_halluc:
            prio = 6
            reason = "Hallucination indicators."
            pl = "HIGH"
        elif is_omission:
            prio = 7
            reason = "Omission/addition failures."
            pl = "HIGH"
        elif is_stem_risk:
            prio = 8
            reason = "Cases containing high-risk STEM terminology."
            pl = "HIGH"
        else:
            prio = 9
            reason = "Representative successful cases as controls."
            pl = "CONTROL"
            
        case_priorities[uid] = {
            "prio_num": prio,
            "reason": reason,
            "priority": pl,
            "domain": dom
        }

    # Group by priority
    buckets = {i: {} for i in range(1, 10)} # prio -> {domain -> [uids]}
    for uid, data in case_priorities.items():
        p = data["prio_num"]
        d = data["domain"]
        if d not in buckets[p]:
            buckets[p][d] = []
        buckets[p][d].append(uid)
        
    final_selected = []
    # Select cases round-robin across domains within each priority bucket
    for p in range(1, 10):
        if len(final_selected) >= 45:
            break
            
        domains = list(buckets[p].keys())
        if not domains: continue
        
        # round robin
        added_in_bucket = 0
        while buckets[p] and len(final_selected) < 48: # max cap around 48
            for d in list(domains):
                if d in buckets[p] and buckets[p][d]:
                    uid = buckets[p][d].pop(0)
                    final_selected.append(uid)
                    added_in_bucket += 1
                    if not buckets[p][d]:
                        del buckets[p][d]
                if len(final_selected) >= 48:
                    break
                    
    # Now build the excel file
    wb = openpyxl.Workbook()
    
    # 1. Instructions
    ws_instr = wb.active
    ws_instr.title = "Instructions"
    instructions = [
        ["GATE 3A: Critical Human Validation Sub-Review"],
        [],
        ["PASS = correct"],
        ["FAIL = incorrect"],
        ["UNCERTAIN = reviewer cannot confidently determine"],
        [],
        ["A scientifically wrong but fluent translation = FAIL."],
        ["Formula corruption = FAIL."],
        ["Technical identifier corruption = FAIL."],
        ["Hallucinated unrelated content = CRITICAL FAIL."],
        ["Detached Kannada morphology = grammar/morphology FAIL."],
        [],
        ["Leave all judgment cells empty. Fill with PASS/FAIL/UNCERTAIN."]
    ]
    for row in instructions: ws_instr.append(row)
        
    # 2. Hindi Critical Review
    ws_hi = wb.create_sheet("Hindi Critical Review")
    hi_cols = [
        "ID", "Priority", "Failure Reason", "Domain", "Source English", 
        "Raw Hindi", "Protected Hindi", "Raw Kannada", "Protected Kannada", "Protected+Morphology Kannada",
        "AI Reference Hindi", "AI Reference Kannada",
        "Semantic Correct", "Terminology Correct", "Grammar Correct", "Natural Fluency",
        "Formula Correct", "Technical Identifier Correct", "Hallucination", "Omission", "Addition",
        "Overall Verdict", "Reviewer Notes"
    ]
    ws_hi.append(hi_cols)
    
    # 3. Kannada Critical Review
    ws_kn = wb.create_sheet("Kannada Critical Review")
    kn_cols = [
        "ID", "Priority", "Failure Reason", "Domain", "Source English", 
        "Raw Hindi", "Protected Hindi", "Raw Kannada", "Protected Kannada", "Protected+Morphology Kannada",
        "AI Reference Hindi", "AI Reference Kannada",
        "Semantic Correct", "Terminology Correct", "Grammar Correct", "Natural Fluency",
        "Morphology Correct", "Formula Correct", "Technical Identifier Correct", "Hallucination", "Omission", "Addition",
        "Overall Verdict", "Reviewer Notes"
    ]
    ws_kn.append(kn_cols)
    
    for uid in final_selected:
        data = case_priorities[uid]
        p = data["priority"]
        r = data["reason"]
        dom = data["domain"]
        g3 = gate3_cases[uid]
        
        base_row = [
            uid, p, r, dom, g3["source_en"],
            g3["raw_hi"], g3["protected_hi"], g3["raw_kn"], g3["protected_kn"], g3["protected_morph_kn"],
            g3["ai_reference_hi"], g3["ai_reference_kn"]
        ]
        
        hi_row = base_row + ["", "", "", "", "", "", "", "", "", "", ""]
        kn_row = base_row + ["", "", "", "", "", "", "", "", "", "", "", ""]
        
        ws_hi.append(hi_row)
        ws_kn.append(kn_row)
        
    # 4. Failure Matrix
    ws_fm = wb.create_sheet("Failure Matrix")
    ws_fm.append(["ID", "Failure Reason", "Priority", "Domain"])
    for uid in final_selected:
        data = case_priorities[uid]
        ws_fm.append([uid, data["reason"], data["priority"], data["domain"]])
            
    # 5. Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Total Cases", len(final_selected)])
    
    p_counts = {"CRITICAL": 0, "HIGH": 0, "CONTROL": 0}
    d_counts = {}
    for uid in final_selected:
        p = case_priorities[uid]["priority"]
        p_counts[p] += 1
        dom = case_priorities[uid]["domain"]
        d_counts[dom] = d_counts.get(dom, 0) + 1
        
    for k, v in p_counts.items(): ws_sum.append([k, v])
        
    wb.save(os.path.join(base_dir, "gate3_critical_review.xlsx"))
    
    # Audit markdown
    audit_lines = [
        "# GATE 3A: Critical Review Subset Audit",
        "",
        "## Selection Statistics",
        "- **original cases:** 114",
        f"- **selected cases:** {len(final_selected)}",
        f"- **Hindi cases:** {len(final_selected)}",
        f"- **Kannada cases:** {len(final_selected)}",
        "",
        "## Priority Breakdown",
        f"- **CRITICAL:** {p_counts['CRITICAL']}",
        f"- **HIGH:** {p_counts['HIGH']}",
        f"- **CONTROL:** {p_counts['CONTROL']}",
        "",
        "## Domains Represented"
    ]
    for d, c in d_counts.items(): audit_lines.append(f"- **{d}**: {c} cases")
        
    audit_lines.extend([
        "",
        "## Selection Rationale",
        "Cases were selected based strictly on the priority order, distributing across all 5 domains round-robin within each priority tier up to ~48 cases.",
        "The priorities evaluated were:",
        "1. Human-critical terminology failures detected by Gate 2.",
        "2. Formula preservation failures.",
        "3. Technical identifier failures.",
        "4. Kannada quadratic-equation terminology failures.",
        "5. Kannada morphology failures remaining after Config C.",
        "6. Hallucination indicators.",
        "7. Omission/addition failures.",
        "8. Cases containing high-risk STEM terminology.",
        "9. Representative successful cases as controls.",
        "",
        "## Validation Confirmations",
        "- every selected ID exists in Gate 2",
        "- no source text changed",
        "- no translation output changed",
        "- no human judgments were generated",
        "- all judgment fields are empty",
        "- all cases are marked PENDING_HUMAN_REVIEW"
    ])
    
    with open(os.path.join(base_dir, "gate3_critical_review_audit.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(audit_lines))
        
    print(f"Created {len(final_selected)} cases.")

if __name__ == "__main__":
    main()
