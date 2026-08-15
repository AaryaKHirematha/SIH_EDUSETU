import os
import openpyxl
from openpyxl.utils import get_column_letter

def main():
    wb_path = "/mnt/d/SIH/gate1_benchmark/gate3_critical_review.xlsx"
    if not os.path.exists(wb_path):
        print(f"Error: {wb_path} does not exist.")
        return

    wb = openpyxl.load_workbook(wb_path)
    
    # Check sheets
    expected_sheets = ["Instructions", "Hindi Critical Review", "Kannada Critical Review"]
    for s in expected_sheets:
        if s not in wb.sheetnames:
            print(f"Error: Missing sheet '{s}'")
            return
            
    # Verify cases and empty fields
    cases_count_hi = 0
    empty_fields_hi = 0
    ws_hi = wb["Hindi Critical Review"]
    hi_headers = [c.value for c in ws_hi[1]]
    judgment_cols_hi = [i for i, h in enumerate(hi_headers) if h in ["Semantic Correct", "Terminology Correct", "Grammar Correct", "Natural Fluency", "Formula Correct", "Technical Identifier Correct", "Hallucination", "Omission", "Addition", "Overall Verdict"]]
    
    for row in ws_hi.iter_rows(min_row=2, max_row=ws_hi.max_row):
        if row[0].value:
            cases_count_hi += 1
            for col_idx in judgment_cols_hi:
                if row[col_idx].value is None or str(row[col_idx].value).strip() == "":
                    empty_fields_hi += 1
                else:
                    print(f"Error: Non-empty judgment found in Hindi sheet, row {row[0].row}, col {col_idx}")

    cases_count_kn = 0
    empty_fields_kn = 0
    ws_kn = wb["Kannada Critical Review"]
    kn_headers = [c.value for c in ws_kn[1]]
    judgment_cols_kn = [i for i, h in enumerate(kn_headers) if h in ["Semantic Correct", "Terminology Correct", "Grammar Correct", "Natural Fluency", "Morphology Correct", "Formula Correct", "Technical Identifier Correct", "Hallucination", "Omission", "Addition", "Overall Verdict"]]

    for row in ws_kn.iter_rows(min_row=2, max_row=ws_kn.max_row):
        if row[0].value:
            cases_count_kn += 1
            for col_idx in judgment_cols_kn:
                if row[col_idx].value is None or str(row[col_idx].value).strip() == "":
                    empty_fields_kn += 1
                else:
                    print(f"Error: Non-empty judgment found in Kannada sheet, row {row[0].row}, col {col_idx}")

    # Verify columns exist (Raw, Protected, Protected+Morphology, English source, AI reference)
    expected_cols_hi = ["Source English", "Raw Hindi", "Protected Hindi", "Raw Kannada", "Protected Kannada", "Protected+Morphology Kannada", "AI Reference Hindi", "AI Reference Kannada"]
    for c in expected_cols_hi:
        if c not in hi_headers: print(f"Missing column in Hindi: {c}")
        if c not in kn_headers: print(f"Missing column in Kannada: {c}")

    # Update Instructions Sheet
    ws_inst = wb["Instructions"]
    
    # Append Gate 3B findings to instructions
    max_r = ws_inst.max_row
    ws_inst.cell(row=max_r + 2, column=1, value="--- GATE 3B MORPHOLOGY UPDATE ---")
    ws_inst.cell(row=max_r + 3, column=1, value="Note to Reviewers (Kannada):")
    ws_inst.cell(row=max_r + 4, column=1, value="1. CS_009 is the only currently identified genuine morphology issue.")
    ws_inst.cell(row=max_r + 5, column=1, value="2. The previous 15 morphology flags were evaluator false positives.")
    ws_inst.cell(row=max_r + 6, column=1, value="3. Human reviewers must judge actual Kannada grammar rather than relying on automated morphology flags.")
    
    # Make them bold
    from openpyxl.styles import Font
    for r in range(max_r + 2, max_r + 7):
        ws_inst.cell(row=r, column=1).font = Font(bold=True)

    wb.save(wb_path)
    
    total_cases = max(cases_count_hi, cases_count_kn)
    total_empty = empty_fields_hi + empty_fields_kn
    
    print("SUCCESS")
    print(f"Total Cases: {total_cases}")
    print(f"Total Blank Fields: {total_empty}")

if __name__ == "__main__":
    main()
